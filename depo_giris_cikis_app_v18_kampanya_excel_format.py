
import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime
from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ============================================================
# PAGE CONFIG
# ============================================================
st.set_page_config(
    page_title="Depo Dashboard",
    layout="wide",
    initial_sidebar_state="expanded"
)

APP_DIR = Path(__file__).parent
HISTORY_DIR = APP_DIR / "history"
HISTORY_DIR.mkdir(exist_ok=True)


# ============================================================
# GENERAL HELPERS
# ============================================================
def clean_text(value):
    if pd.isna(value):
        return ""

    text = str(value).strip().lower()

    replacements = {
        "ı": "i",
        "İ": "i",
        "ğ": "g",
        "ü": "u",
        "ş": "s",
        "ö": "o",
        "ç": "c",
        "â": "a",
        "î": "i",
        "û": "u",
    }

    for old, new in replacements.items():
        text = text.replace(old, new)

    text = text.replace("-", " ").replace("_", " ").replace("/", " ")
    text = " ".join(text.split())

    return text


def normalize_code(value):
    """Ürün kodlarını 504.0 / '00504' / '4188 UN1266' gibi durumlara karşı normalize eder."""
    if pd.isna(value):
        return None

    text = str(value).strip()

    match = re.search(r"\d+", text)
    if match:
        text = match.group(0)

    if text.endswith(".0"):
        text = text[:-2]

    return text.lstrip("0") or "0"


def excel_serial_to_date(value):
    """Excel serial, string veya datetime değerini tarihe çevirir."""
    if pd.isna(value):
        return pd.NaT

    if isinstance(value, (int, float, np.integer, np.floating)):
        # Excel serial date
        if value > 10000:
            return pd.to_datetime("1899-12-30") + pd.to_timedelta(float(value), unit="D")

    return pd.to_datetime(value, dayfirst=True, errors="coerce")


def looks_like_date(value):
    return not pd.isna(excel_serial_to_date(value))


def get_week_start(date_series):
    date_series = pd.to_datetime(date_series, errors="coerce")
    return date_series - pd.to_timedelta(date_series.dt.weekday, unit="D")


def safe_format_cell(x):
    if isinstance(x, (int, float, np.integer, np.floating)) and not pd.isna(x):
        return f"{x:,.0f}"
    return x


def safe_divide(a, b):
    if b == 0:
        return 0
    return a / b


def add_months(ts, months):
    return pd.Timestamp(ts) + pd.DateOffset(months=months)


# ============================================================
# AUTH
# ============================================================
def get_secret_value(key, default_value):
    try:
        return st.secrets[key]
    except Exception:
        return default_value


def login():
    planning_password = get_secret_value("PLANNING_PASSWORD", "planning2026")
    depo_password = get_secret_value("DEPO_PASSWORD", "depo2026")

    if "role" not in st.session_state:
        st.session_state["role"] = None

    if st.session_state["role"]:
        return st.session_state["role"]

    logo_path = APP_DIR / "logo.png"
    if logo_path.exists():
        st.image(str(logo_path), width=110)

    st.title("Depo Dashboard")
    st.caption("Lütfen size verilen şifre ile giriş yapın.")

    password = st.text_input("Şifre", type="password")

    if st.button("Giriş Yap"):
        if password == planning_password:
            st.session_state["role"] = "planning"
            st.rerun()
        elif password == depo_password:
            st.session_state["role"] = "depo"
            st.rerun()
        else:
            st.error("Şifre hatalı.")

    st.stop()


# ============================================================
# FILE READERS
# ============================================================
def find_product_code_column(df):
    candidates = []

    for col in df.columns:
        col_text = clean_text(col)
        if (
            "product" in col_text or
            "code" in col_text or
            "kod" in col_text or
            "article" in col_text or
            "current" in col_text
        ):
            candidates.append(col)

    for col in candidates:
        if df[col].notna().sum() > 0:
            return col

    for col in df.columns:
        if df[col].notna().sum() > 0:
            return col

    return df.columns[0]


def find_product_name_column(df, code_col):
    for col in df.columns:
        col_text = clean_text(col)
        if col != code_col and ("name" in col_text or "libelle" in col_text or "urun adi" in col_text):
            return col

    cols = list(df.columns)
    if code_col in cols:
        idx = cols.index(code_col)
        if idx + 1 < len(cols):
            return cols[idx + 1]

    return code_col


def normalize_product_type(value):
    text = clean_text(value)

    if not text:
        return "Ana Ürün"

    mini_keywords = ["mini", "sample", "tester", "sachet", "numune", "deneme"]
    if any(k in text for k in mini_keywords):
        return "Mini Sample"

    ana_keywords = ["ana", "main", "regular", "standart", "standard"]
    if any(k in text for k in ana_keywords):
        return "Ana Ürün"

    return "Ana Ürün"


def read_code_sheet(xls, possible_sheet_names):
    existing_sheets = {clean_text(s): s for s in xls.sheet_names}

    selected_sheet = None
    for name in possible_sheet_names:
        key = clean_text(name)
        if key in existing_sheets:
            selected_sheet = existing_sheets[key]
            break

    if selected_sheet is None:
        return set(), None

    df = pd.read_excel(xls, sheet_name=selected_sheet)
    df = df.dropna(how="all")

    if df.empty:
        return set(), selected_sheet

    first_valid_col = None
    for col in df.columns:
        if df[col].notna().sum() > 0:
            first_valid_col = col
            break

    if first_valid_col is None:
        return set(), selected_sheet

    codes = set(df[first_valid_col].apply(normalize_code).dropna().astype(str))
    return codes, selected_sheet


def read_campaign_sheet(xls):
    campaign_sheet = None

    for sheet in xls.sheet_names:
        sheet_key = clean_text(sheet)
        if "kampanya" in sheet_key or "campaign" in sheet_key:
            campaign_sheet = sheet
            break

    if campaign_sheet is None:
        return pd.DataFrame(columns=["campaign", "start", "end"])

    raw = pd.read_excel(xls, sheet_name=campaign_sheet, header=None)
    raw = raw.dropna(how="all")

    if raw.empty:
        return pd.DataFrame(columns=["campaign", "start", "end"])

    header_row = None
    campaign_col = None
    start_col = None
    end_col = None

    for idx in range(min(15, len(raw))):
        row = raw.iloc[idx]

        for col_idx, value in row.items():
            text = clean_text(value)
            original = str(value).lower()

            if not text:
                continue

            if campaign_col is None and (
                "window" in text or "campaign" in text or "kampanya" in text or "period" in text
            ):
                campaign_col = col_idx

            if start_col is None and (
                text == "start" or "baslangic" in text or "başlangıç" in original or "start date" in text
            ):
                start_col = col_idx

            if end_col is None and (
                text == "end" or "bitis" in text or "bitiş" in original or "end date" in text
            ):
                end_col = col_idx

        if campaign_col is not None and start_col is not None and end_col is not None:
            header_row = idx
            break

    if header_row is None:
        header_row = 2 if len(raw) > 2 else 0
        campaign_col = 0
        start_col = 2
        end_col = 3

    data = raw.iloc[header_row + 1:].copy()
    if data.empty:
        data = raw.iloc[header_row:].copy()

    campaign = pd.DataFrame({
        "campaign": data.iloc[:, campaign_col] if campaign_col < data.shape[1] else "",
        "start": data.iloc[:, start_col] if start_col < data.shape[1] else pd.NaT,
        "end": data.iloc[:, end_col] if end_col < data.shape[1] else pd.NaT,
    })

    campaign = campaign.dropna(subset=["campaign"])
    campaign["campaign"] = campaign["campaign"].astype(str).str.strip()
    campaign = campaign[campaign["campaign"] != ""]

    campaign["start"] = campaign["start"].apply(excel_serial_to_date)
    campaign["end"] = campaign["end"].apply(excel_serial_to_date)

    campaign = campaign.dropna(subset=["start", "end"])
    campaign = campaign[campaign["end"] >= campaign["start"]]

    return campaign[["campaign", "start", "end"]].reset_index(drop=True)


def read_mapping_file(mapping_file):
    xls = pd.ExcelFile(mapping_file)

    ana_codes, ana_sheet = read_code_sheet(
        xls,
        ["Ana Ürün", "Ana Urun", "Ana", "Main", "Ana Kodlar"]
    )

    mini_codes, mini_sheet = read_code_sheet(
        xls,
        ["Mini Sample", "Mini", "Sample", "Mini Kodlar", "Sample Kodlar"]
    )

    adr_codes, adr_sheet = read_code_sheet(
        xls,
        ["ADR", "Adr", "ADR Kodlar", "Adr Kodlar"]
    )

    mapping_rows = []

    for code in ana_codes:
        mapping_rows.append({"product_code": code, "product_type": "Ana Ürün"})

    for code in mini_codes:
        mapping_rows.append({"product_code": code, "product_type": "Mini Sample"})

    type_map = pd.DataFrame(mapping_rows)

    # Eski ürün tipi sheet varsa eksikleri tamamla
    if any(clean_text(s) == clean_text("ürün tipi") for s in xls.sheet_names):
        product_type_df = pd.read_excel(mapping_file, sheet_name="ürün tipi")
        product_type_df.columns = [str(c).strip() for c in product_type_df.columns]

        code_col_candidates = ["Current Code", "Kod", "Code", "Product Code", "Ürün Kodu", "Urun Kodu"]
        type_col_candidates = ["Ürün Tipi", "Urun Tipi", "Product Type", "Tip"]

        code_col = next((c for c in code_col_candidates if c in product_type_df.columns), None)
        type_col = next((c for c in type_col_candidates if c in product_type_df.columns), None)

        if code_col and type_col:
            product_type_df["product_code"] = product_type_df[code_col].apply(normalize_code)
            product_type_df["product_type"] = product_type_df[type_col].apply(normalize_product_type)

            old_map = (
                product_type_df[["product_code", "product_type"]]
                .dropna()
                .drop_duplicates(subset=["product_code"], keep="first")
            )

            type_map = pd.concat([old_map, type_map], ignore_index=True)
            type_map = type_map.drop_duplicates(subset=["product_code"], keep="last")

    if type_map.empty:
        type_map = pd.DataFrame(columns=["product_code", "product_type"])
    else:
        type_map = type_map.drop_duplicates(subset=["product_code"], keep="last")

    campaign = read_campaign_sheet(xls)

    sheet_info = {
        "Ana Ürün Sheet": ana_sheet or "-",
        "Mini Sample Sheet": mini_sheet or "-",
        "ADR Sheet": adr_sheet or "-",
        "Ana Kod Sayısı": len(ana_codes),
        "Mini Kod Sayısı": len(mini_codes),
        "ADR Kod Sayısı": len(adr_codes),
    }

    return type_map, adr_codes, campaign, sheet_info


def read_supply_file(supply_file):
    xls = pd.ExcelFile(supply_file)

    sheet_name = None
    for s in xls.sheet_names:
        if clean_text(s) == "supply":
            sheet_name = s
            break
    if sheet_name is None:
        sheet_name = xls.sheet_names[0]

    raw = pd.read_excel(supply_file, sheet_name=sheet_name, header=None)

    header_row = None
    for i in range(min(10, len(raw))):
        row = raw.iloc[i]
        date_count = sum(looks_like_date(v) for v in row.tolist())
        if date_count >= 2:
            header_row = i
            break

    if header_row is None:
        header_row = 0

    headers = list(raw.iloc[header_row])
    df = raw.iloc[header_row + 1:].copy()
    df.columns = headers
    df = df.dropna(how="all")
    df = df.loc[:, df.columns.notna()]

    code_col = find_product_code_column(df)
    name_col = find_product_name_column(df, code_col)

    date_cols = [col for col in df.columns if col not in [code_col, name_col] and looks_like_date(col)]
    if not date_cols:
        date_cols = [col for col in df.columns if col not in [code_col, name_col]]

    df = df.rename(columns={code_col: "product_code", name_col: "product_name"})
    df["product_code"] = df["product_code"].apply(normalize_code)

    long_df = df.melt(
        id_vars=["product_code", "product_name"],
        value_vars=date_cols,
        var_name="date",
        value_name="inbound_qty"
    )

    long_df["date"] = long_df["date"].apply(excel_serial_to_date)
    long_df["date"] = long_df["date"] - pd.Timedelta(days=7)
    long_df["inbound_qty"] = pd.to_numeric(long_df["inbound_qty"], errors="coerce").fillna(0)

    long_df = long_df.dropna(subset=["product_code", "date"])
    long_df = long_df[long_df["inbound_qty"] != 0]
    long_df["week_start"] = get_week_start(long_df["date"])

    return long_df


def read_apo_file(apo_file):
    xls = pd.ExcelFile(apo_file)

    sheet_name = None
    for s in xls.sheet_names:
        if "weekly" in clean_text(s) and "forecast" in clean_text(s):
            sheet_name = s
            break
    if sheet_name is None:
        sheet_name = xls.sheet_names[0]

    raw = pd.read_excel(apo_file, sheet_name=sheet_name, header=None)

    date_row = None
    for i in range(min(10, len(raw))):
        row = raw.iloc[i]
        date_count = sum(looks_like_date(v) for v in row.tolist())
        if date_count >= 2:
            date_row = i
            break

    if date_row is None:
        date_row = 1 if len(raw) > 1 else 0

    date_values = list(raw.iloc[date_row, 1:])

    df = raw.iloc[date_row + 1:].copy()
    df = df.dropna(how="all")

    code_col = df.columns[0]
    qty_cols = list(df.columns[1:])

    rename_map = {code_col: "product_code"}
    for col, date_value in zip(qty_cols, date_values):
        rename_map[col] = excel_serial_to_date(date_value)

    df = df.rename(columns=rename_map)
    date_cols = [c for c in df.columns if isinstance(c, pd.Timestamp)]

    if not date_cols:
        date_cols = [c for c in df.columns if c != "product_code"]

    df["product_code"] = df["product_code"].apply(normalize_code)

    long_df = df.melt(
        id_vars=["product_code"],
        value_vars=date_cols,
        var_name="date",
        value_name="outbound_qty"
    )

    long_df["date"] = pd.to_datetime(long_df["date"], errors="coerce")
    long_df["outbound_qty"] = pd.to_numeric(long_df["outbound_qty"], errors="coerce").fillna(0)

    long_df = long_df.dropna(subset=["product_code", "date"])
    long_df = long_df[long_df["outbound_qty"] != 0]
    long_df["week_start"] = get_week_start(long_df["date"])

    return long_df


def read_ekol_file(ekol_file):
    """
    Ekol dosyasını esnek okur.
    Hem haftalık depo yeri stoklarını hem de kapasite özetini ayrı tablolar olarak çıkarmaya çalışır.
    """
    if ekol_file is None:
        return None, None

    xls = pd.ExcelFile(ekol_file)
    sheet_name = xls.sheet_names[0]
    raw = pd.read_excel(ekol_file, sheet_name=sheet_name, header=None)

    # Kapasite özetini bulmaya çalış
    capacity_rows = []
    for i in range(len(raw)):
        row_text = " ".join([str(x) for x in raw.iloc[i].dropna().tolist()])
        if "Kapasite" in row_text or "Doluluk" in row_text or "Boş" in row_text or "Bos" in row_text:
            # Bu satırdan sonraki 10 satırı dene
            cap = raw.iloc[i:i+12].copy()
            cap = cap.dropna(how="all")
            if not cap.empty:
                capacity_rows = cap
            break

    ekol_capacity = None
    if isinstance(capacity_rows, pd.DataFrame) and not capacity_rows.empty:
        ekol_capacity = capacity_rows.reset_index(drop=True)

    # Haftalık stok tablosu: ilk iki kolon + tarih kolonları gibi oku
    header_row = None
    for i in range(min(10, len(raw))):
        row = raw.iloc[i]
        date_count = sum(looks_like_date(v) for v in row.tolist())
        if date_count >= 2:
            header_row = i
            break

    ekol_weekly = None
    if header_row is not None:
        headers = list(raw.iloc[header_row])
        df = raw.iloc[header_row + 1:].copy()
        df.columns = headers
        df = df.dropna(how="all")
        ekol_weekly = df

    return ekol_weekly, ekol_capacity


# ============================================================
# CALCULATION
# ============================================================
def add_product_type(df, type_map, adr_codes):
    df = df.copy()
    df["product_code"] = df["product_code"].apply(normalize_code)

    df = df.merge(type_map, on="product_code", how="left")
    df["product_type"] = df["product_type"].fillna("Ana Ürün")
    df["product_type"] = df["product_type"].replace({"Merch": "Ana Ürün", "Diğer": "Ana Ürün"})
    df["is_adr"] = df["product_code"].isin(adr_codes)

    # ADR ana ürün içinde kalır, ayrıca ek hesap olarak gösterilir.
    df["report_type"] = df["product_type"]

    return df


def assign_campaign(week_start, campaign_df):
    if campaign_df.empty or pd.isna(week_start):
        return ""

    week_start = pd.to_datetime(week_start)
    week_end = week_start + pd.Timedelta(days=6)

    active = campaign_df[
        (campaign_df["start"] <= week_end) &
        (campaign_df["end"] >= week_start)
    ]

    if active.empty:
        return ""

    return " / ".join(active["campaign"].astype(str).unique())


def prepare_report(
    supply_file,
    apo_file,
    mapping_file,
    ana_palet_ici,
    mini_palet_ici,
    adr_palet_ici,
    sarf_palet,
    tir_kapasitesi,
    initial_ana,
    initial_mini,
    initial_adr,
    depo_kapasitesi,
    takip_esigi,
    kritik_esigi
):
    type_map, adr_codes, campaign_df, sheet_info = read_mapping_file(mapping_file)

    supply_long = read_supply_file(supply_file)
    apo_long = read_apo_file(apo_file)

    supply_long = add_product_type(supply_long, type_map, adr_codes)
    apo_long = add_product_type(apo_long, type_map, adr_codes)

    # Ana/Mini hesapları: ADR ana içinde kalır
    inbound_base = (
        supply_long
        .groupby(["week_start", "report_type"], as_index=False)["inbound_qty"]
        .sum()
    )

    outbound_base = (
        apo_long
        .groupby(["week_start", "report_type"], as_index=False)["outbound_qty"]
        .sum()
    )

    # ADR ayrıca gösterilir
    inbound_adr = (
        supply_long[supply_long["is_adr"]]
        .groupby("week_start", as_index=False)["inbound_qty"]
        .sum()
    )
    inbound_adr["report_type"] = "ADR"

    outbound_adr = (
        apo_long[apo_long["is_adr"]]
        .groupby("week_start", as_index=False)["outbound_qty"]
        .sum()
    )
    outbound_adr["report_type"] = "ADR"

    inbound = pd.concat([inbound_base, inbound_adr], ignore_index=True)
    outbound = pd.concat([outbound_base, outbound_adr], ignore_index=True)

    movement = pd.merge(
        inbound,
        outbound,
        on=["week_start", "report_type"],
        how="outer"
    ).fillna(0)

    all_weeks = pd.DataFrame({"week_start": sorted(movement["week_start"].dropna().unique())})
    all_types = pd.DataFrame({"report_type": ["Ana Ürün", "Mini Sample", "ADR"]})
    grid = all_weeks.merge(all_types, how="cross")

    movement = grid.merge(movement, on=["week_start", "report_type"], how="left").fillna(0)
    movement = movement.sort_values(["report_type", "week_start"])

    pallet_map = {
        "Ana Ürün": ana_palet_ici,
        "Mini Sample": mini_palet_ici,
        "ADR": adr_palet_ici,
    }

    initial_stock_map = {
        "Ana Ürün": initial_ana,
        "Mini Sample": initial_mini,
        "ADR": initial_adr,
    }

    stock_rows = []

    for report_type, g in movement.groupby("report_type"):
        current_stock = initial_stock_map.get(report_type, 0)

        for _, row in g.sort_values("week_start").iterrows():
            current_stock = current_stock + row["inbound_qty"] - row["outbound_qty"]

            pallet_inner = pallet_map.get(report_type, ana_palet_ici)
            pallet = safe_divide(current_stock, pallet_inner)

            stock_rows.append({
                "week_start": row["week_start"],
                "report_type": report_type,
                "inbound_qty": row["inbound_qty"],
                "outbound_qty": row["outbound_qty"],
                "stock_qty": current_stock,
                "pallet_inner": pallet_inner,
                "pallet": pallet,
            })

    detail = pd.DataFrame(stock_rows)

    weekly = (
        detail
        .pivot_table(
            index="week_start",
            columns="report_type",
            values=["inbound_qty", "outbound_qty", "stock_qty", "pallet"],
            aggfunc="sum"
        )
    )

    weekly.columns = [f"{metric}_{rtype}" for metric, rtype in weekly.columns]
    weekly = weekly.reset_index()

    # Streamlit Cloud / farklı pandas versiyonlarında week_start bazen PeriodIndex kalabiliyor.
    # Bu nedenle rapor tarih alanlarına geçmeden önce güvenli datetime formatına çeviriyoruz.
    weekly["week_start"] = pd.to_datetime(weekly["week_start"].astype(str), errors="coerce")

    needed_cols = [
        "inbound_qty_Ana Ürün", "outbound_qty_Ana Ürün", "stock_qty_Ana Ürün", "pallet_Ana Ürün",
        "inbound_qty_Mini Sample", "outbound_qty_Mini Sample", "stock_qty_Mini Sample", "pallet_Mini Sample",
        "inbound_qty_ADR", "outbound_qty_ADR", "stock_qty_ADR", "pallet_ADR",
    ]

    for col in needed_cols:
        if col not in weekly.columns:
            weekly[col] = 0

    report = pd.DataFrame()

    # Cloud ortamında week_start bazen Period/obj olarak gelebiliyor.
    # Bu yüzden strftime öncesinde kesin datetime'a çeviriyoruz.
    weekly["week_start"] = pd.to_datetime(weekly["week_start"], errors="coerce")

    weekly["week_start"] = pd.to_datetime(weekly["week_start"], errors="coerce")
    report["Hafta"] = weekly["week_start"].dt.strftime("%Y-W%U")
    report["Hafta Başlangıcı"] = weekly["week_start"].dt.strftime("%d.%m.%Y")
    report["Kampanya"] = weekly["week_start"].apply(lambda x: assign_campaign(x, campaign_df))

    report["Ana Ürün Giriş"] = weekly["inbound_qty_Ana Ürün"]
    report["Ana Ürün Çıkış"] = weekly["outbound_qty_Ana Ürün"]
    report["Ana Ürün Ekol Stok Seviyesi"] = weekly["stock_qty_Ana Ürün"]
    report["Ana Ürün Palet"] = weekly["pallet_Ana Ürün"]
    report["Ana Ürün Giriş Paleti"] = report["Ana Ürün Giriş"] / ana_palet_ici

    report["Mini Sample Giriş"] = weekly["inbound_qty_Mini Sample"]
    report["Mini Sample Çıkış"] = weekly["outbound_qty_Mini Sample"]
    report["Mini Sample Ekol Stok Seviyesi"] = weekly["stock_qty_Mini Sample"]
    report["Mini Sample Palet"] = weekly["pallet_Mini Sample"]
    report["Mini Sample Giriş Paleti"] = report["Mini Sample Giriş"] / mini_palet_ici

    report["ADR Giriş"] = weekly["inbound_qty_ADR"]
    report["ADR Çıkış"] = weekly["outbound_qty_ADR"]
    report["ADR Ekol Stok Seviyesi"] = weekly["stock_qty_ADR"]
    report["ADR Palet"] = weekly["pallet_ADR"]

    report["Sarf Palet"] = sarf_palet
    report["ADR Düşülecek Palet"] = report["ADR Palet"]
    report["Total Palet"] = (
        report["Ana Ürün Palet"] +
        report["Mini Sample Palet"] +
        report["Sarf Palet"] -
        report["ADR Düşülecek Palet"]
    )

    # Tır: sadece o haftaki giriş paleti
    report["Tır Sayısı"] = (
        report["Ana Ürün Giriş Paleti"] +
        report["Mini Sample Giriş Paleti"]
    ) / tir_kapasitesi

    report["Kapasite Kullanım %"] = report["Total Palet"] / depo_kapasitesi * 100
    report["Kalan Kapasite Palet"] = depo_kapasitesi - report["Total Palet"]

    report["Haftalık Palet Değişimi"] = report["Total Palet"].diff().fillna(0)

    def capacity_status(value):
        if value >= 100:
            return "Kapasite Aşımı"
        if value >= kritik_esigi:
            return "Kritik"
        if value >= takip_esigi:
            return "Takip"
        return "Güvenli"

    report["Kapasite Durumu"] = report["Kapasite Kullanım %"].apply(capacity_status)
    report["Palet Trend"] = np.where(
        report["Haftalık Palet Değişimi"] > 0,
        "Artış",
        np.where(report["Haftalık Palet Değişimi"] < 0, "Azalış", "Sabit")
    )

    numeric_cols = report.select_dtypes(include=[np.number]).columns
    report[numeric_cols] = report[numeric_cols].round(0)

    # Aylık özet
    monthly_summary = movement.copy()
    monthly_summary["month"] = pd.to_datetime(monthly_summary["week_start"]).dt.strftime("%Y-%m")
    monthly_summary = (
        monthly_summary
        .groupby(["month", "report_type"], as_index=False)[["inbound_qty", "outbound_qty"]]
        .sum()
    )

    monthly_summary = monthly_summary.pivot_table(
        index="month",
        columns="report_type",
        values=["inbound_qty", "outbound_qty"],
        aggfunc="sum",
        fill_value=0
    )

    monthly_summary.columns = [f"{metric}_{rtype}" for metric, rtype in monthly_summary.columns]
    monthly_summary = monthly_summary.reset_index()

    for col in [
        "inbound_qty_Ana Ürün", "outbound_qty_Ana Ürün",
        "inbound_qty_Mini Sample", "outbound_qty_Mini Sample",
        "inbound_qty_ADR", "outbound_qty_ADR"
    ]:
        if col not in monthly_summary.columns:
            monthly_summary[col] = 0

    monthly_report = pd.DataFrame()
    monthly_report["Ay"] = monthly_summary["month"]
    monthly_report["Ana Ürün Giriş"] = monthly_summary["inbound_qty_Ana Ürün"]
    monthly_report["Ana Ürün Çıkış"] = monthly_summary["outbound_qty_Ana Ürün"]
    monthly_report["Mini Sample Giriş"] = monthly_summary["inbound_qty_Mini Sample"]
    monthly_report["Mini Sample Çıkış"] = monthly_summary["outbound_qty_Mini Sample"]
    monthly_report["ADR Giriş"] = monthly_summary["inbound_qty_ADR"]
    monthly_report["ADR Çıkış"] = monthly_summary["outbound_qty_ADR"]

    monthly_numeric_cols = monthly_report.select_dtypes(include=[np.number]).columns
    monthly_report[monthly_numeric_cols] = monthly_report[monthly_numeric_cols].round(0)

    # Mevcut hafta palet tablosu
    mevcut_ana_palet = safe_divide(initial_ana, ana_palet_ici)
    mevcut_mini_palet = safe_divide(initial_mini, mini_palet_ici)
    mevcut_adr_palet = safe_divide(initial_adr, adr_palet_ici)
    mevcut_total_palet = mevcut_ana_palet + mevcut_mini_palet + sarf_palet - mevcut_adr_palet

    mevcut_hafta_report = pd.DataFrame({
        "Kategori": ["Ana Ürün", "Mini Sample", "ADR", "Sarf", "Total"],
        "Başlangıç Stok": [initial_ana, initial_mini, initial_adr, np.nan, np.nan],
        "Palet İçi": [ana_palet_ici, mini_palet_ici, adr_palet_ici, np.nan, np.nan],
        "Mevcut Hafta Palet": [
            mevcut_ana_palet,
            mevcut_mini_palet,
            mevcut_adr_palet,
            sarf_palet,
            mevcut_total_palet
        ]
    })

    for col in ["Başlangıç Stok", "Palet İçi", "Mevcut Hafta Palet"]:
        mevcut_hafta_report[col] = pd.to_numeric(mevcut_hafta_report[col], errors="coerce").round(0)

    # Kategori kontrol
    supply_check_base = supply_long.copy()
    supply_check_adr = supply_long[supply_long["is_adr"]].copy()
    supply_check_adr["report_type"] = "ADR"
    supply_check_all = pd.concat([supply_check_base, supply_check_adr], ignore_index=True)

    apo_check_base = apo_long.copy()
    apo_check_adr = apo_long[apo_long["is_adr"]].copy()
    apo_check_adr["report_type"] = "ADR"
    apo_check_all = pd.concat([apo_check_base, apo_check_adr], ignore_index=True)

    supply_category_check = (
        supply_check_all
        .groupby("report_type")
        .agg(
            supply_total_qty=("inbound_qty", "sum"),
            supply_product_count=("product_code", "nunique")
        )
        .reset_index()
    )

    apo_category_check = (
        apo_check_all
        .groupby("report_type")
        .agg(
            apo_total_qty=("outbound_qty", "sum"),
            apo_product_count=("product_code", "nunique")
        )
        .reset_index()
    )

    category_check = pd.merge(
        supply_category_check,
        apo_category_check,
        on="report_type",
        how="outer"
    ).fillna(0)

    return {
        "report": report,
        "monthly_report": monthly_report,
        "mevcut_hafta_report": mevcut_hafta_report,
        "detail": detail,
        "campaign_df": campaign_df,
        "sheet_info": sheet_info,
        "category_check": category_check,
    }


# ============================================================
# STYLING
# ============================================================
def highlight_increased_pallet_columns(dataframe, increased_weeks):
    styles = pd.DataFrame("", index=dataframe.index, columns=dataframe.columns)

    for col in dataframe.columns:
        week_key = str(col).split("\n")[0]
        if week_key in increased_weeks:
            styles[col] = "background-color: #f8d7da; color: #842029; font-weight: 600;"

    return styles


def highlight_after_horizon_columns(dataframe, horizon_week_keys):
    styles = pd.DataFrame("", index=dataframe.index, columns=dataframe.columns)

    for col in dataframe.columns:
        week_key = str(col).split("\n")[0]
        if week_key in horizon_week_keys:
            styles[col] = "background-color: #fde2e2; border-left: 4px solid #c0392b;"

    return styles


def highlight_capacity_and_kpi(df, takip_esigi=85, kritik_esigi=99):
    styles = pd.DataFrame("", index=df.index, columns=df.columns)

    for row_label in df.index:
        row_text = str(row_label)

        if "Palet" in row_text:
            styles.loc[row_label, :] = "background-color: #d6eaf8;"

        if "ADR Palet" in row_text:
            styles.loc[row_label, :] = "background-color: #fdebd0;"

        if "Total Palet" in row_text:
            styles.loc[row_label, :] = "background-color: #d5f5e3; font-weight: bold;"

        if "Tır Sayısı" in row_text:
            styles.loc[row_label, :] = "background-color: #e8daef; font-weight: bold;"

        if "Kapasite Kullanım %" in row_text:
            for col in df.columns:
                val = pd.to_numeric(pd.Series([df.loc[row_label, col]]), errors="coerce").iloc[0]
                if pd.isna(val):
                    continue
                if val >= 100:
                    styles.loc[row_label, col] = "background-color: #c0392b; color: white; font-weight: bold;"
                elif val >= kritik_esigi:
                    styles.loc[row_label, col] = "background-color: #f5b7b1; font-weight: bold;"
                elif val >= takip_esigi:
                    styles.loc[row_label, col] = "background-color: #f9e79f; font-weight: bold;"
                else:
                    styles.loc[row_label, col] = "background-color: #abebc6; font-weight: bold;"

    return styles


def format_excel_workbook(excel_bytes):
    bio = BytesIO(excel_bytes)
    wb = load_workbook(bio)

    header_fill = PatternFill("solid", fgColor="EAF2F8")
    header_font = Font(bold=True)
    number_format = '#,##0'

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = number_format

        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                value = cell.value
                if value is not None:
                    max_len = max(max_len, len(str(value)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 30)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


# ============================================================
# SCREEN HELPERS
# ============================================================
def show_header(title, subtitle=None):
    logo_path = APP_DIR / "logo.png"

    if logo_path.exists():
        col_logo, col_title = st.columns([1, 8])
        with col_logo:
            st.image(str(logo_path), width=85)
        with col_title:
            st.title(title)
            if subtitle:
                st.caption(subtitle)
    else:
        st.title(title)
        if subtitle:
            st.caption(subtitle)


def render_planning_screen():
    show_header("Depo Kapasite Dashboard", "Planning ekranı")

    with st.sidebar:
        if st.button("Çıkış Yap"):
            st.session_state["role"] = None
            st.rerun()

        st.header("Dosyalar")
        supply_file = st.file_uploader("Supply dosyası", type=["xlsx", "xlsm"], key="planning_supply")
        apo_file = st.file_uploader("APO Forecast dosyası", type=["xlsx", "xlsm"], key="planning_apo")
        mapping_file = st.file_uploader("Ürün Tipi + Kampanya + ADR dosyası", type=["xlsx", "xlsm"], key="planning_mapping")
        ekol_file = st.file_uploader("Ekol Depo Data Dosyası", type=["xlsx", "xlsm"], key="planning_ekol")

        st.header("Palet Parametreleri")
        ana_palet_ici = st.number_input("Ana Ürün Palet İçi", min_value=1, value=2400, step=50)
        mini_palet_ici = st.number_input("Mini Sample Palet İçi", min_value=1, value=15000, step=100)
        adr_palet_ici = st.number_input("ADR Palet İçi", min_value=1, value=5540, step=10)
        sarf_palet = st.number_input("Haftalık Sarf Palet", min_value=0.0, value=250.0, step=0.5)
        tir_kapasitesi = st.number_input("1 Tır Kaç Palet?", min_value=1, value=40, step=1)

        st.header("Başlangıç Stok")
        initial_ana = st.number_input("Başlangıç Ana Ürün Stok", value=0, step=1000)
        initial_mini = st.number_input("Başlangıç Mini Sample Stok", value=0, step=1000)
        initial_adr = st.number_input("Başlangıç ADR Stok", value=0, step=1000)

        st.header("Depo Kapasite Parametreleri")
        depo_kapasitesi = st.number_input("Depo Maksimum Palet Kapasitesi", min_value=1.0, value=1100.0, step=100.0)
        takip_esigi = st.number_input("Takip Eşiği (%)", min_value=0.0, max_value=100.0, value=85.0, step=1.0)
        kritik_esigi = st.number_input("Kritik Eşiği (%)", min_value=0.0, max_value=100.0, value=99.0, step=1.0)

        calculate = st.button("Raporu Hesapla", type="primary")

    if calculate:
        if not supply_file or not apo_file or not mapping_file:
            st.error("Lütfen Supply, APO Forecast ve Ürün Tipi dosyalarının üçünü de yükleyin.")
            st.stop()

        with st.spinner("Rapor hazırlanıyor..."):
            data = prepare_report(
                supply_file=supply_file,
                apo_file=apo_file,
                mapping_file=mapping_file,
                ana_palet_ici=ana_palet_ici,
                mini_palet_ici=mini_palet_ici,
                adr_palet_ici=adr_palet_ici,
                sarf_palet=sarf_palet,
                tir_kapasitesi=tir_kapasitesi,
                initial_ana=initial_ana,
                initial_mini=initial_mini,
                initial_adr=initial_adr,
                depo_kapasitesi=depo_kapasitesi,
                takip_esigi=takip_esigi,
                kritik_esigi=kritik_esigi
            )

        report = data["report"]
        monthly_report = data["monthly_report"]
        mevcut_hafta_report = data["mevcut_hafta_report"]
        detail = data["detail"]
        campaign_df = data["campaign_df"]
        sheet_info = data["sheet_info"]
        category_check = data["category_check"]

        # KPI'lar ilk hafta
        first_total = report["Total Palet"].iloc[0]
        first_capacity = report["Kapasite Kullanım %"].iloc[0]
        first_remaining = report["Kalan Kapasite Palet"].iloc[0]
        first_status = report["Kapasite Durumu"].iloc[0]
        first_week = report["Hafta"].iloc[0]

        # Peak ilk 5 ay
        weekly_for_peak = report.copy()
        weekly_for_peak["_date"] = pd.to_datetime(weekly_for_peak["Hafta Başlangıcı"], dayfirst=True, errors="coerce")
        first_date = weekly_for_peak["_date"].min()
        horizon_date = add_months(first_date, 5)
        first_5_months = weekly_for_peak[weekly_for_peak["_date"] < horizon_date].copy()
        if first_5_months.empty:
            first_5_months = weekly_for_peak.copy()

        peak_idx = first_5_months["Total Palet"].idxmax()
        peak_week = first_5_months.loc[peak_idx, "Hafta"]
        peak_pallet = first_5_months.loc[peak_idx, "Total Palet"]
        increasing_week_count = int((first_5_months["Palet Trend"] == "Artış").sum())

        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("İlk Hafta Total Palet", f"{first_total:,.0f}", first_week)
        c2.metric("İlk Hafta Kapasite %", f"{first_capacity:,.0f}%")
        c3.metric("İlk Hafta Kalan Kapasite", f"{first_remaining:,.0f} palet")
        c4.metric("Peak Hafta / İlk 5 Ay", f"{peak_week}", f"{peak_pallet:,.0f} palet")
        c5.metric("Artan Hafta Sayısı / İlk 5 Ay", f"{increasing_week_count}")

        if first_status in ["Kritik", "Kapasite Aşımı"]:
            st.error(f"İlk hafta kapasite durumu: {first_status}")
        elif first_status == "Takip":
            st.warning(f"İlk hafta kapasite durumu: {first_status}")
        else:
            st.success(f"İlk hafta kapasite durumu: {first_status}")

        # Yatay tablo
        st.subheader("Yatay Haftalık Görünüm / Excel Formatı")

        pivot_rows = [
            "Ana Ürün Giriş",
            "Ana Ürün Çıkış",
            "Ana Ürün Ekol Stok Seviyesi",
            "Ana Ürün Palet",

            "Mini Sample Giriş",
            "Mini Sample Çıkış",
            "Mini Sample Ekol Stok Seviyesi",
            "Mini Sample Palet",

            "ADR Giriş",
            "ADR Çıkış",
            "ADR Ekol Stok Seviyesi",
            "ADR Palet",

            "Sarf Palet",
            "ADR Düşülecek Palet",
            "Total Palet",
            "Kapasite Kullanım %",
            "Kalan Kapasite Palet",
            "Kapasite Durumu",
            "Palet Trend",
            "Tır Sayısı",
        ]

        horizontal = report.set_index("Hafta")[pivot_rows].T

        weekly_date_map = report.set_index("Hafta")["Hafta Başlangıcı"]
        campaign_row = report.set_index("Hafta")["Kampanya"]

        horizontal.columns = [
            f"{week}\n{weekly_date_map.loc[week]}\n{campaign_row.loc[week] if campaign_row.loc[week] else ''}"
            for week in horizontal.columns
        ]

        weekly_sorted = report.copy()
        weekly_sorted["Total Palet Artış"] = weekly_sorted["Total Palet"].diff()
        increased_weeks = weekly_sorted.loc[
            weekly_sorted["Total Palet Artış"] > 0, "Hafta"
        ].astype(str).tolist()

        weekly_horizon = report.copy()
        weekly_horizon["_date"] = pd.to_datetime(weekly_horizon["Hafta Başlangıcı"], dayfirst=True, errors="coerce")
        first_week_date = weekly_horizon["_date"].min()
        horizon_limit = add_months(first_week_date, 5)
        after_horizon_weeks = weekly_horizon.loc[
            weekly_horizon["_date"] >= horizon_limit, "Hafta"
        ].astype(str).tolist()

        st.caption("Kırmızı kolonlar: Total Palet bir önceki haftaya göre artan haftaları gösterir. Açık kırmızı/sol çizgili kolonlar: ilk 5 aydan sonrası için data eksik olabilir; bu alan tam doğru sonucu vermeyebilir.")

        styled_horizontal = (
            horizontal
            .style
            .apply(lambda _: highlight_increased_pallet_columns(horizontal, increased_weeks), axis=None)
            .apply(lambda _: highlight_after_horizon_columns(horizontal, after_horizon_weeks), axis=None)
            .apply(lambda _: highlight_capacity_and_kpi(horizontal, takip_esigi, kritik_esigi), axis=None)
            .format(safe_format_cell, na_rep="")
        )

        st.dataframe(styled_horizontal, use_container_width=True)

        with st.expander("Okunan Kampanya Takvimi"):
            if campaign_df.empty:
                st.warning("Kampanya takvimi okunamadı. Kampanya sheetinde kampanya adı, start ve end tarihleri olduğundan emin olun.")
            else:
                campaign_view = campaign_df.copy()
                campaign_view["start"] = campaign_view["start"].dt.strftime("%d.%m.%Y")
                campaign_view["end"] = campaign_view["end"].dt.strftime("%d.%m.%Y")
                st.dataframe(campaign_view, use_container_width=True)

        st.subheader("Mevcut Hafta Palet Tablosu")
        st.dataframe(
            mevcut_hafta_report.style.format({
                "Başlangıç Stok": lambda x: "" if pd.isna(x) else f"{x:,.0f}",
                "Palet İçi": lambda x: "" if pd.isna(x) else f"{x:,.0f}",
                "Mevcut Hafta Palet": lambda x: "" if pd.isna(x) else f"{x:,.0f}",
            }),
            use_container_width=True
        )

        st.subheader("Aylık Giriş / Çıkış Toplamları")
        monthly_horizontal = monthly_report.set_index("Ay").T
        st.dataframe(
            monthly_horizontal.style.format(safe_format_cell, na_rep=""),
            use_container_width=True
        )

        # Ekol data
        if ekol_file is not None:
            st.subheader("Ekol Depo Data")
            ekol_weekly, ekol_capacity = read_ekol_file(ekol_file)

            if ekol_capacity is not None:
                st.write("Ekol Kapasite Özeti")
                st.dataframe(ekol_capacity, use_container_width=True)

            if ekol_weekly is not None:
                st.write("Ekol Haftalık Stok / Doluluk Tablosu")
                st.dataframe(ekol_weekly, use_container_width=True)

        # Depo ekranının göreceği güvenli operasyon sheet'i
        depo_export_cols = [
            "Hafta",
            "Hafta Başlangıcı",
            "Kampanya",
            "Ana Ürün Giriş",
            "Ana Ürün Çıkış",
            "Mini Sample Giriş",
            "Mini Sample Çıkış",
            "ADR Giriş",
            "ADR Çıkış",
        ]
        depo_export = report[depo_export_cols].copy()

        # Export
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            horizontal.to_excel(writer, sheet_name="Yatay Özet")
            depo_export.to_excel(writer, sheet_name="Depo Operasyon", index=False)
            mevcut_hafta_report.to_excel(writer, sheet_name="Mevcut Hafta Palet", index=False)
            monthly_horizontal.to_excel(writer, sheet_name="Aylık Giriş Çıkış")
            report.to_excel(writer, sheet_name="Veri", index=False)
            pd.DataFrame([sheet_info]).to_excel(writer, sheet_name="Okunan Sheet Bilgisi", index=False)
            category_check.to_excel(writer, sheet_name="Kategori Kontrol", index=False)
            campaign_df.to_excel(writer, sheet_name="Kampanya", index=False)

            # Ekol dosyası yüklendiyse geçmiş rapora da ekle
            if ekol_file is not None:
                ekol_weekly_export, ekol_capacity_export = read_ekol_file(ekol_file)

                if ekol_capacity_export is not None:
                    ekol_capacity_export.to_excel(writer, sheet_name="Ekol Kapasite Özeti", index=False, header=False)

                if ekol_weekly_export is not None:
                    ekol_weekly_export.to_excel(writer, sheet_name="Ekol Haftalık Stok", index=False)

        formatted_report_bytes = format_excel_workbook(output.getvalue())
        st.session_state["last_report_bytes"] = formatted_report_bytes
        st.session_state["last_report_default_name"] = f"depo_giris_cikis_raporu_{datetime.now().strftime('%Y%m%d')}"

        st.download_button(
            label="Excel Raporu İndir",
            data=st.session_state["last_report_bytes"],
            file_name="depo_giris_cikis_dashboard_raporu.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # Save panel
    st.divider()
    st.subheader("Raporu Geçmişe Kaydet")

    if "last_report_bytes" in st.session_state:
        report_save_name = st.text_input(
            "Geçmişe kaydetme adı",
            value=st.session_state.get("last_report_default_name", f"depo_giris_cikis_raporu_{datetime.now().strftime('%Y%m%d')}")
        )

        if st.button("Geçmişe Kaydet"):
            clean_name = "".join(
                ch if ch.isalnum() or ch in [" ", "_", "-"] else "_"
                for ch in report_save_name.strip()
            )

            if not clean_name:
                clean_name = f"depo_giris_cikis_raporu_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

            history_filename = f"{clean_name}.xlsx"
            history_path = HISTORY_DIR / history_filename

            if history_path.exists():
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                history_filename = f"{clean_name}_{timestamp}.xlsx"
                history_path = HISTORY_DIR / history_filename

            history_path.write_bytes(st.session_state["last_report_bytes"])
            st.success(f"Rapor geçmişe kaydedildi: {history_filename}")
    else:
        st.info("Önce raporu hesaplayın. Rapor hesaplandıktan sonra burada geçmişe kaydedebilirsiniz.")

    # History
    st.divider()
    st.subheader("Geçmiş Raporlar")

    history_files = sorted(HISTORY_DIR.glob("*.xlsx"), reverse=True)

    if history_files:
        selected_history = st.selectbox(
            "Geçmiş rapor seç",
            options=[f.name for f in history_files]
        )

        selected_path = HISTORY_DIR / selected_history

        with open(selected_path, "rb") as f:
            st.download_button(
                "Seçili Geçmiş Raporu İndir",
                data=f.read(),
                file_name=selected_history,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        if st.button("Seçili Geçmiş Raporu Aç"):
            xls_history = pd.ExcelFile(selected_path)
            tabs = st.tabs(xls_history.sheet_names)

            for tab, sheet_name in zip(tabs, xls_history.sheet_names):
                with tab:
                    old_df = pd.read_excel(selected_path, sheet_name=sheet_name)
                    st.dataframe(old_df, use_container_width=True)

        st.caption(f"Toplam kayıtlı rapor sayısı: {len(history_files)}")
    else:
        st.info("Henüz geçmiş rapor yok. Raporu oluşturduktan sonra 'Geçmişe Kaydet' butonuna basarsan burada görünür.")


def render_depo_screen():
    show_header("Depo Operasyon Ekranı", "Kayıtlı operasyon raporları")

    with st.sidebar:
        if st.button("Çıkış Yap"):
            st.session_state["role"] = None
            st.rerun()

    st.info("Planning ekranında kaydedilen raporları buradan görüntüleyebilirsiniz.")

    history_files = sorted(HISTORY_DIR.glob("*.xlsx"), reverse=True)

    if not history_files:
        st.warning("Henüz kayıtlı rapor bulunmuyor. Önce Planning ekranından rapor oluşturup 'Geçmişe Kaydet' yapmalısınız.")
        return

    selected_history = st.selectbox(
        "Görüntülenecek raporu seç",
        options=[f.name for f in history_files]
    )

    selected_path = HISTORY_DIR / selected_history

    try:
        xls_history = pd.ExcelFile(selected_path)
    except Exception as e:
        st.error(f"Rapor açılırken hata oluştu: {e}")
        return

    allowed_sheets = [
        "Depo Operasyon",
        "Aylık Giriş Çıkış",
        "Kampanya",
        "Ekol Kapasite Özeti",
        "Ekol Haftalık Stok",
    ]

    available_allowed_sheets = [s for s in allowed_sheets if s in xls_history.sheet_names]

    if not available_allowed_sheets:
        st.error("Bu kayıtlı raporda depo ekranı için uygun sheet bulunamadı. Raporu güncel kodla tekrar kaydedin.")
        return

    tabs = st.tabs(available_allowed_sheets)

    for tab, sheet_name in zip(tabs, available_allowed_sheets):
        with tab:
            df = pd.read_excel(selected_path, sheet_name=sheet_name)

            # Planlama hesaplarını gizle; Ekol sheetleri orijinal depo verisi olduğu için dokunma.
            if sheet_name not in ["Ekol Kapasite Özeti", "Ekol Haftalık Stok"]:
                hidden_keywords = [
                    "Palet",
                    "Tır",
                    "Kapasite",
                    "Stok Seviyesi",
                    "Düşülecek",
                    "Trend",
                ]

                visible_cols = [
                    col for col in df.columns
                    if not any(keyword.lower() in str(col).lower() for keyword in hidden_keywords)
                ]

                df = df[visible_cols]

            st.dataframe(
                df.style.format(safe_format_cell, na_rep=""),
                use_container_width=True
            )

    with open(selected_path, "rb") as f:
        st.download_button(
            "Seçili Raporu İndir",
            data=f.read(),
            file_name=selected_history,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


# ============================================================
# MAIN
# ============================================================
role = login()

if role == "planning":
    render_planning_screen()
elif role == "depo":
    render_depo_screen()
